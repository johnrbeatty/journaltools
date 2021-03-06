import argparse
import os
import csv

from openpyxl import Workbook
from journaltools import dateconvert

# This is an alternate routine to dc-convert.py & journaltools.convertcsv that is used with CSV files exported
# from mdgen-blr.py. It converts and exports those files to an Excel file that can be cut and pasted into a
# Digital Commons batch import spreadsheet.


def convertcsv(input_file, output_file, verbose):
    # Import a CSV file that has been exported from this code. Read file, everything except the start and end PDF
    # pages, then spit them out into an Excel workbook.

    # Set Excel workbook filename
    wb = Workbook()
    ws = wb.active
    wbrow = 1
    dest_filename = output_file + '.xlsx'

    if verbose:
        print(f'Reading {input_file}')

    # Declare/clear variables
    title = ""
    volume = ''
    start_page = ""
    issue = ''
    month = ''
    year = ''
    document_type = ''
    f_name1 = ""
    m_name1 = ""
    l_name1 = ""
    suffix1 = ""
    f_name2 = ""
    m_name2 = ""
    l_name2 = ""
    suffix2 = ""
    f_name3 = ""
    m_name3 = ""
    l_name3 = ""
    suffix3 = ""
    f_name4 = ""
    m_name4 = ""
    l_name4 = ""
    suffix4 = ""

    # Step through each row of input CSV file. Read in each row and assign to variables. Write to Excel file.
    # The columns are hard-coded to correspond to the Digital Commons import columns for Buffalo Law Review.
    # To be more useful, this could write to an output and write the values in the columns based on the
    # headers.
    with open(input_file, newline='') as csvfile:
        data_reader = csv.reader(csvfile, delimiter=",", quotechar="'")
        for row in data_reader:
            try:
                title = row[0]
                volume = row[1]
                start_page = row[2]
                issue = row[3]
                month = row[4]
                year = row[5]
                document_type = row[6]
                f_name1 = (row[7])
                m_name1 = (row[8])
                l_name1 = (row[9])
                suffix1 = (row[10])
                f_name2 = (row[11])
                m_name2 = (row[12])
                l_name2 = (row[13])
                suffix2 = (row[14])
                f_name3 = (row[15])
                m_name3 = (row[16])
                l_name3 = (row[17])
                suffix3 = (row[18])
                f_name4 = (row[19])
                m_name4 = (row[20])
                l_name4 = (row[21])
                suffix4 = (row[22])
            except IndexError:
                print('Warning. List out of range. Check export.')

            date = dateconvert(year, month)

            ws.cell(row=wbrow, column=1, value=title)
            ws.cell(row=wbrow, column=42, value=volume)
            ws.cell(row=wbrow, column=37, value=start_page)
            ws.cell(row=wbrow, column=36, value=issue)
            ws.cell(row=wbrow, column=39, value=date)
            ws.cell(row=wbrow, column=35, value=document_type)
            ws.cell(row=wbrow, column=5, value=f_name1)
            ws.cell(row=wbrow, column=6, value=m_name1)
            ws.cell(row=wbrow, column=7, value=l_name1)
            ws.cell(row=wbrow, column=8, value=suffix1)
            if f_name1 != '' and l_name1 == '':
                ws.cell(row=wbrow, column=11, value='TRUE')
            else:
                ws.cell(row=wbrow, column=11, value='FALSE')
            ws.cell(row=wbrow, column=12, value=f_name2)
            ws.cell(row=wbrow, column=13, value=m_name2)
            ws.cell(row=wbrow, column=14, value=l_name2)
            ws.cell(row=wbrow, column=15, value=suffix2)
            if f_name2 != '' and l_name2 == '':
                ws.cell(row=wbrow, column=18, value='TRUE')
            else:
                ws.cell(row=wbrow, column=18, value='FALSE')
            ws.cell(row=wbrow, column=19, value=f_name3)
            ws.cell(row=wbrow, column=20, value=m_name3)
            ws.cell(row=wbrow, column=21, value=l_name3)
            ws.cell(row=wbrow, column=22, value=suffix3)
            if f_name3 != '' and l_name3 == '':
                ws.cell(row=wbrow, column=25, value='TRUE')
            else:
                ws.cell(row=wbrow, column=25, value='FALSE')
            ws.cell(row=wbrow, column=26, value=f_name4)
            ws.cell(row=wbrow, column=27, value=m_name4)
            ws.cell(row=wbrow, column=28, value=l_name4)
            ws.cell(row=wbrow, column=29, value=suffix4)
            if f_name4 != '' and l_name4 == '':
                ws.cell(row=wbrow, column=32, value='TRUE')
            else:
                ws.cell(row=wbrow, column=32, value='FALSE')
            wbrow += 1

    csvfile.close()
    wb.save(filename=dest_filename)
    if verbose:
        print(f'Saving {dest_filename}')
    wb.close()


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('-v', '--verbose',
                        action='store_true',
                        dest='verbose',
                        help='Print status messages.',
                        )
    parser.add_argument('-o', '--output-file',
                        dest='destination',
                        type=str,
                        help='Output file. Default is <input file>.xslx',
                        )
    parser.add_argument('input_file',
                        type=str,
                        help="Import CSV file to be used for PDF splitting. Must be in same format as export.")
    args = parser.parse_args()

    # Split filename from extension before passing to the various functions. Use input filename for template
    # if no output filename specified.
    if args.destination:
        output_file, output_extension = os.path.splitext(args.destination)
    else:
        output_file, output_extension = os.path.splitext(args.input_file)

    # Read input_file and export to Excel
    convertcsv(args.input_file, output_file, args.verbose)


if __name__ == '__main__':
    main()
