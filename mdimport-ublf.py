import os
import argparse

from openpyxl import load_workbook
from journaltools import exportcsvnew
from journaltools import capitalize_title

# This program takes in a user-compiled Excel file with the journal metadata and converts it to a CSV file for use
# with the other tools here. It's used in place of dsplit for journals which are unsuitable for automatic
# metadata collecting.
#
# This is set up for the UB Law Forum. One of its quirks is that most articles are in sections. There is no
# metadata for the sections in the repository, so the code here prepends the section to the article title.
# A better option is probably to figure out a cleaner way to represent the metadata in the repository.


def importxl(import_file):
    # Import CSV file for this issue. User should in theory use a template with the right fields. Note to self: create
    # a template. Read each row, do some minor processing, then put the contents into a set of variables.
    # Eventually,  return the variables so they can be passed to another procedure that will write a CSV file that
    # can be used to split the full issue PDF and also to be converted to Digital Commons format for "easy" batch
    # importing. Har har har.

    title = []
    page = []
    pdf_start_page = []
    pdf_end_page = []
    author = []
    section = []

    # Set defaults for columns; will be overwritten as necessary
    section_col = 1
    title_col = 2
    page_col = 3
    start_col = 4
    end_col = 5
    first_col = 6
    middle_col = 7
    last_col = 8
    suffix_col = 9
    author2_first_col = 0
    author2_middle_col = 0
    author2_last_col = 0
    author2_suffix_col = 0

    wb = load_workbook(filename=import_file, data_only=True)
    ws = wb.active
    # Read first row and get headers.
    headers = []
    for c in range(1, ws.max_column+1):
        headers.append(ws.cell(row=1, column=c).internal_value)
    for c in range(0, len(headers)):
        if headers[c] == 'section':
            section_col = c + 1
        if headers[c] == 'title':
            title_col = c + 1
        if headers[c] == 'page':
            page_col = c + 1
        if headers[c] == 'start_pdf_page':
            start_col = c + 1
        if headers[c] == 'end_pdf_page':
            end_col = c + 1
        if headers[c] == 'author_first':
            first_col = c + 1
        if headers[c] == 'author_middle':
            middle_col = c + 1
        if headers[c] == 'author_last':
            last_col = c + 1
        if headers[c] == 'author_suffix':
            suffix_col = c + 1
        if headers[c] == 'author2_first':
            author2_first_col = c + 1
        if headers[c] == 'author2_middle':
            author2_middle_col = c + 1
        if headers[c] == 'author2_last':
            author2_last_col = c + 1
        if headers[c] == 'author2_suffix':
            author2_suffix_col = c + 1

    # Iterate through all rows, reading values into lists to pass back to main.
    max_row = ws.max_row
    for i in range(2, max_row+1):
        section_temp = ws.cell(row=i, column=section_col).internal_value
        if section_temp:
            section_temp = section_temp.title()
            section_temp = capitalize_title(section_temp)
        section.append(section_temp)
        temp_title = ws.cell(row=i, column=title_col).internal_value
        temp_title = temp_title.title()
        temp_title = capitalize_title(temp_title)
        title.append(temp_title)
        page_temp = ws.cell(row=i, column=page_col).internal_value
        if page_temp:
            page.append(page_temp)
        else:
            page.append('')
        pdf_start_page.append(ws.cell(row=i, column=start_col).internal_value)
        pdf_end_page.append(ws.cell(row=i, column=end_col).internal_value)
        author_temp = ws.cell(row=i, column=first_col).internal_value, \
            ws.cell(row=i, column=middle_col).internal_value, ws.cell(row=i, column=last_col).internal_value, \
            ws.cell(row=i, column=suffix_col).internal_value
        author_list = [author_temp]
        # Only look for second author if there were second author columns in the input file.
        if author2_first_col:
            # If there are columns in the input file for a second author, check to make sure there's a value in the
            # first name field. If there is, then pull all four columns into a tuple, then append it to the list.
            if ws.cell(row=i, column=author2_first_col).internal_value:
                author_temp = ws.cell(row=i, column=author2_first_col).internal_value, \
                               ws.cell(row=i, column=author2_middle_col).internal_value, \
                               ws.cell(row=i, column=author2_last_col).internal_value, \
                               ws.cell(row=i, column=author2_suffix_col).internal_value
                author_list.append(author_temp)

        if author_list:
            author.append(author_list)
        else:
            author.append('')

    return title, page, pdf_start_page, pdf_end_page, author, section


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument('filename',
                        help='PDF file to process.',
                        type=str,
                        )
    parser.add_argument('-v', '--verbose',
                        action='store_true',
                        dest='verbose',
                        help='Print status messages.',
                        )
    parser.add_argument('-t', '--test',
                        action='store_true',
                        help="Test only. Don't output any files. Use with debug options to see test output.",
                        )
    parser.add_argument('-d', '--debug',
                        action='store_true',
                        dest='debug',
                        help="Show debug output.",
                        default=0,
                        )
    parser.add_argument('-o', '--output-file',
                        dest='destination',
                        type=str,
                        help='Use supplied filename as filename template for output files.',
                        )
    args = parser.parse_args()

    # Split filename from extension before passing to the various functions. Use input filename for template
    # if no output filename specified.
    if args.destination:
        output_file, output_extension = os.path.splitext(args.destination)
    else:
        output_file, output_extension = os.path.splitext(args.filename)

    # Fetch metadata from import Excel file
    title, start_page, start_pdf_page, end_pdf_page, author, section = importxl(args.filename)
    # Export CSV file, or show what output would be if test flag is set
    exportcsvnew(output_file, args.verbose, args.debug, args.test, title, start_page, start_pdf_page,
                 end_pdf_page, author, section=section)


if __name__ == '__main__':
    main()
