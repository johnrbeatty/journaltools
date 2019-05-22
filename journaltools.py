import fnmatch
import os
import re
import io
import csv

import PyPDF2

from openpyxl import Workbook
from openpyxl import load_workbook
from pdfminer.pdfinterp import PDFResourceManager, PDFPageInterpreter
from pdfminer.converter import TextConverter
from pdfminer.layout import LAParams
from pdfminer.pdfpage import PDFPage


def splitname(full_name):
    # Take in author's full name and attempt to split into first, middle, last, suffix and return all.

    m_name = ''
    l_name = ''
    suffix = ''

    # Split names with a comma. This will catch many names with a suffix (e.g., James Q. Author, Jr.). If there is
    # anything after a comma, capitalize it (this is included to fix names in all caps) and assign it to suffix.
    split_names = re.split(r', ', full_name)
    if len(split_names) > 1:
        suffix = split_names[1].capitalize()

    # Split the full name into every part separated by a space. Assume the first part is the first name and assign
    # it to f_name.
    split_names = re.split(r'\s{1,2}', split_names[0])
    f_name = split_names[0].capitalize()

    # Check to see how many parts the name was split into. If it's four, assume that the last part is a two word last
    # name. Assign the second part to m_name and the last two parts to l_name. No this doesn't deal well with people
    # with three word last names, thank you for asking. Also, it doesn't deal well with people with two middle names,
    # but then neither does the repository that all this stuff is going into. These just have to be fixed by hand.
    if len(split_names) == 4:
        m_name = split_names[1].capitalize()
        l_name = split_names[2].capitalize() + ' ' + split_names[3].capitalize()
    # If there are three names, then assign the second one to m_name and the third to l_name.
    elif len(split_names) == 3:
        m_name = split_names[1].capitalize()
        l_name = split_names[2].capitalize()
    # IF there are only two names, there is likely not a middle name, so assign the second name to l_name.
    elif len(split_names) == 2:
        l_name = split_names[1].capitalize()
    return f_name, m_name, l_name, suffix


def splitpdf(filename, verbose, debug, start_pdf_page, end_pdf_page, output_file):
    # Take input of a source PDF filename, verbose and debug flags, lists of starting and ending pages of individual
    # articles in a single PDF file, and an output file name template. Open the input file, and then create individual
    # files for each page range represented in the two lists.

    if verbose:
        print("Exporting split PDFs:")
    # Check to make sure that there are the same number of start and end pages. If not, give the user a warning.
    if len(start_pdf_page) != len(end_pdf_page):
        print('Missing page number. Check input file.')
    # Loop through a range of numbers between 0 and the number of start pages. For each start page value, open the
    # input file, pull out the pages between the start page and the matching end page, and write them to a new file
    # with the output file name, plus the number of the loop counter variable.
    for r in range(0, len(start_pdf_page)):
        # This debug output is probably mostly useful if the OCR on your PDF files is good enough to do the metadata
        # gathering and PDF splitting in one step. None of ours were, so the pages were all checked by hand before
        # splitting. If yours are better, this could be useful while you're tweaking the regular expressions to extract
        # your metadata.
        if 2 < debug < 5:
            print(f'Record: {r}')
            print(f'Start page: {start_pdf_page[r]}')
            print(f'End page: {end_pdf_page[r]}')

        export_file = output_file + '-' + str(r) + '.pdf'
        if verbose:
            print(f'Exporting to {export_file}')
        pdf_output_file = open(export_file, 'wb')
        pdf_writer = PyPDF2.PdfFileMerger()
        # Check to make sure that the page ranges make sense. Is the end page after the start page? If so, copy pages
        # to a new file. If not, give the user an error message.
        if end_pdf_page > start_pdf_page:
            pdf_writer.append(filename, pages=(start_pdf_page[r], end_pdf_page[r]+1))
        else:
            print(f'End page ({end_pdf_page[r]}) is after start page ({start_pdf_page[r]}) for record {r}')
        pdf_writer.write(pdf_output_file)
        pdf_output_file.close()


def getpdf(filename, maxpages, verbose, debug):
    # This code is a little ugly. There is no real documentation for PDFMiner. But I used because its text extraction
    # is better than PyPDF2. Because of the lack of documentation, I had to work from someone else's demo code to
    # start. I really should fix the code to allow the user to pass arguments to the converter. It will, for example,
    # decrypt a password-protected file if the user can send the password.

    # Take in a filename and verbose and debug flags. Open the PDF file filename, then step through each page,
    # fetch the OCR text, and then add it to a list. When the last page is read, close the file and return the list.
    # Yes, this probably uses a lot of memory and is inefficient. But given that I needed multiple command-line
    # front-ends that each looked for slightly different metadata in slightly different places, it ended up being much
    # easier to separate this code entirely, read the entire file in at once, and pass it all back to the calling
    # function. Pulling in one page at a time for analysis was messy. Maybe there's a better way to do this, but I
    # haven't had any issues yet processing PDFs up to about 10MB/150 pages.

    # open file
    pdf_file_obj = open(filename, 'rb')

    # open TextConverter device from pdfminer; there should probably be code in here to pass arguments
    # other than the filename to the converter. Maxpages has been added. Password still to be added.
    # FIXME: Add password functionality. Check to see if there are any other arguments that should be passed.
    rsrcmgr = PDFResourceManager()
    retstr = io.StringIO()
    codec = 'utf-8'
    laparams = LAParams()
    device = TextConverter(rsrcmgr, retstr, codec=codec, laparams=laparams)
    interpreter = PDFPageInterpreter(rsrcmgr, device)
    password = ""
    caching = True
    pagenos = set()

    # Create list  for OCR text from PDF pages
    page_text = []

    if verbose:
        print(f'Reading file: {filename}\n')

    # Step through each page in PDF from first page (page 0) to page maxpages (default=0/all), append all text on
    # page to page_text. Then close file and return page_text.
    for pageNumber, page in enumerate(PDFPage.get_pages(pdf_file_obj, pagenos, maxpages=maxpages, password=password,
                                                        caching=caching, check_extractable=True)):
        # fetch OCR page text; print page text at debug levels 4 & 5
        interpreter.process_page(page)
        page_text.append(retstr.getvalue())
        if 3 < debug < 6:
            print(f'Page number {pageNumber}:')
            print(page_text[pageNumber])
            print('\n')

        # Clear retstr and move current position to start. This clears the last read page from memory before
        # grabbing the next page.
        retstr.truncate(0)
        retstr.seek(0)

    pdf_file_obj.close()
    device.close()
    retstr.close()

    return page_text


def exportcsvnew(output_file, verbose, debug, test, title, start_page, start_pdf_page, end_pdf_page, author,
                 section=None):
    # All new export routine that uses tuples for the author names. Most of the metadata scraping routines need to be
    # updated to use this routine instead of the old one. The old code uses two lists for the authors, author 1
    # and author 2. author1 was a list for the first (or sole) author on each piece and author2 was a list for
    # the second author (if any) on each piece. The name was stored as a string and split using the splitnames
    # function. The new code uses one list (author) that contains nested lists of tuples. Each author name is split on
    # intake into four parts and put into a tuple. If there are multiple authors on a piece, they are all put into
    # a list and that list is entered as a list item in the author list.
    #
    # This takes a filename, creates that file and then basically just dumps all of the passed lists into the file.
    # It also writes column headers, which are now necessary because the import routines check the headers.

    # Export collected metadata to CSV file. If test flag set, display metadata instead.
    if not test:
        headers = []
        author_count = 0
        # Build the headers.
        if title:
            headers.append('title')
        if section:
            headers.append('section')
        if start_page:
            headers.append('start_page')
        if start_pdf_page:
            headers.append('start_pdf_page')
        if end_pdf_page:
            headers.append('end_pdf_page')
        for c in range(0, len(author)):
            if len(author[c]) > author_count:
                author_count = len(author[c])
        for c in range(0, author_count):
            headers.append('f_name' + str(c+1))
            headers.append('m_name' + str(c+1))
            headers.append('l_name' + str(c+1))
            headers.append('suffix' + str(c+1))

        export_file = output_file + '.csv'
        with open(export_file, 'w', newline='\n') as csvfile:
            data_writer = csv.writer(
                csvfile,
                delimiter=',',
                quotechar='"',
                doublequote=True,
                escapechar=" ",
                quoting=csv.QUOTE_MINIMAL)
            # Write column headers
            data_writer.writerow(headers)
            # Step through each record and write to CSV
            for r in range(0, len(title)):
                # Check author list to make sure there is data for four authors. For any list entry with no data,
                # append a blank string. This will avoid errors and write a blank entry to the csv file.
                for c in range(0, 4):
                    if len(author[r]) < c+1:
                        author[r].append(('', '', '', ''))
                # Build row
                row_data = []
                if title:
                    row_data.append(title[r])
                if section:
                    row_data.append(section[r])
                if start_page:
                    row_data.append(start_page[r])
                if start_pdf_page:
                    row_data.append(start_pdf_page[r])
                if end_pdf_page:
                    row_data.append(end_pdf_page[r])
                if author:
                    for c in range(0, len(author[r])):
                        for d in range(0, 4):
                            row_data.append(author[r][c][d])

                if debug:
                    print(row_data)
                data_writer.writerow(row_data)
            csvfile.close()
            if verbose:
                print("Data written to file: %s" % export_file)
    else:
        # Show test export data. Test is currently broken while I figure out a more efficient way to make it work.
        print("\nTest export data:")
        for r in range(0, len(title)):
            # Unpack authors
            print(author[r])
            author_first, author_middle, author_last, author_suffix = author[r]
            print(f'{title[r]}, {start_page[r]}, {start_pdf_page[r]}, {end_pdf_page[r]}, {author_first},'
                  f' {author_middle}, {author_last}, {author_suffix}')


def importcsv(filename, debug):
    # This allows two-stage metadata-gathering and PDF splitting. By default, the dsplit code will search the input
    # PDF for metadata, write a CSV, then immediately split PDFs. If you have good OCR, that might work most of the
    # time. Where it doesn't, this routine will allow the user to hand-correct the CSV file before splitting the
    # input PDF.

    # Import a CSV file that has been exported from this code. Read file, import all the values for start_pdf_page and
    # end_pdf_page as two lists, then return for use by splitpdf.
    start_pdf_page = []
    end_pdf_page = []
    start_col = 0
    end_col = 0

    # TODO: Add verbose option?
    with open(filename, newline='') as csvfile:
        data_reader = csv.reader(csvfile, delimiter=",", quotechar="'")
        for row in data_reader:
            if start_col == 0 and end_col == 0:
                for c in range(0, len(row)):
                    if row[c] == "start_pdf_page":
                        start_col = c
                    if row[c] == "end_pdf_page":
                        end_col = c
            try:
                start_pdf_page.append(int(row[start_col]))
                end_pdf_page.append(int(row[end_col]))
            except ValueError:
                print('Start or ending PDF pages appear to be missing. Check input file.')
    csvfile.close()
    if debug:
        print(start_pdf_page)
        print(end_pdf_page)
    return start_pdf_page, end_pdf_page


def doublepages(input_file, output_file, verbose):
    # This will double any page wider than 700 points in a PDF. It is the first part of a workflow designed to
    # crop files that were scanned from stapled magazine pages, with only the cover cropped. The routine checks
    # the page width. If narrowed than 700 points, it assigns a value of "S" to the page_type list. If it's wider,
    # it copies the page and assigns "L" as the page type to one and "R" to the other copy. This tells croppages
    # which side of the full page to keep, left or right. "S" pages are cropped to 8.5 x 11.

    page_type = []
    if verbose:
        print(f'Processing {input_file}')
    pdf_output_file = open(output_file, 'wb')
    pdf_writer = PyPDF2.PdfFileWriter()
    input_pdf = open(input_file, 'rb')
    pdf_reader = PyPDF2.PdfFileReader(input_pdf, strict=False)

    # Step through file and double all but the first and last pages
    max_pages = pdf_reader.getNumPages()
    for page_number in range(0, max_pages):
        page_obj = pdf_reader.getPage(page_number)
        # Check page size
        if page_obj.mediaBox.upperRight[0] < 700:
            # single page
            pdf_writer.addPage(page_obj)
            page_type.append('S')
        else:
            # double page
            pdf_writer.addPage(page_obj)
            page_type.append('L')
            pdf_writer.addPage(page_obj)
            page_type.append('R')

    if verbose:
        print('Writing temp file')
    pdf_writer.write(pdf_output_file)
    input_pdf.close()
    pdf_output_file.close()

    return page_type


def croppages(input_file, output_file, verbose, debug):
    # Set temporary filename and call doublepages. This will make a copy of all large pages that need to be split, and
    # assign a type (S=Single page, regular crop; L=Large page, left side crop; R=Large page, right side crop).
    temp_filename, temp_extension = os.path.splitext(input_file)
    temp_filename = temp_filename + '-temp' + temp_extension
    page_type = doublepages(input_file, temp_filename, verbose)
    # TODO: Integrate doublepages into this code?

    if verbose:
        print(f'Processing {temp_filename}')
    pdf_output_file = open(output_file, 'wb')
    pdf_writer = PyPDF2.PdfFileWriter()
    input_pdf = open(temp_filename, 'rb')
    pdf_reader = PyPDF2.PdfFileReader(input_pdf, strict=False)

    # Step through file and crop pages
    max_pages = pdf_reader.getNumPages()
    for page_number in range(0, max_pages):
        page_obj = pdf_reader.getPage(page_number)
        lower_left_x, lower_left_y = page_obj.mediaBox.lowerLeft
        upper_right_x, upper_right_y = page_obj.mediaBox.upperRight

        # Debugging statements
        if debug:
            print(page_number)
            print(lower_left_x, lower_left_y, upper_right_x, upper_right_y)
            print(page_obj.mediaBox.lowerLeft)
            print(page_obj.mediaBox.lowerRight)
            print(page_obj.mediaBox.upperLeft)
            print(page_obj.mediaBox.upperRight)

        # Check page size and crop
        if page_type[page_number] == 'S':
            # single page
            page_obj.mediaBox.lowerLeft = (upper_right_x - 612, upper_right_y - 792)
            pdf_writer.addPage(page_obj)
        elif page_type[page_number] == 'L':
            # double page; left crop
            page_obj.mediaBox.lowerLeft = (0, upper_right_y - 792)
            page_obj.mediaBox.upperRight = (612, upper_right_y)
            pdf_writer.addPage(page_obj)
        elif page_type[page_number] == 'R':
            # double page; right crop
            page_obj.mediaBox.lowerLeft = (upper_right_x - 612, upper_right_y - 792)
            page_obj.mediaBox.upperRight = (upper_right_x, upper_right_y)
            pdf_writer.addPage(page_obj)

    pdf_writer.write(pdf_output_file)
    input_pdf.close()
    pdf_output_file.close()
    if verbose:
        print('Deleting temp file')
    if os.path.exists(temp_filename):
        os.remove(temp_filename)


def combinepdf(input_files, output_file, verbose):
    # This will quickly combine several PDFs into one file. It can be used with bash scripts to combine things that
    # are regularly split up in a file set. I used this to recombine parts of indexes for a law journal into
    # one file for each volume.

    # Set and open output file.
    if verbose:
        print(f'Combining {input_files}')
    pdf_output_file = open(output_file, 'wb')
    pdf_writer = PyPDF2.PdfFileMerger()

    # Step through files in input_files list and combine them into output_file
    for r in range(0, len(input_files)):
        pdf_writer.append(input_files[r])
    pdf_writer.write(pdf_output_file)
    pdf_output_file.close()


def shiftpage(input_file1, input_file2, output_file, verbose):
    # Take input of two file names. Fetch the last page from the second file and add it to the end of the first file.
    # This is necessary for journal articles that were split improperly because the second article starts on the same
    # page as the end of the first article. This will duplicate that page so that the user doesn't have to also
    # download the second article to get the last page of the first.

    if verbose:
        print(f'Opening {input_file1}')
    pdf_output_file = open(output_file, 'wb')
    pdf_writer = PyPDF2.PdfFileMerger()

    # Use PdfFileMerger to append the first page of input_file2 to the end of input_file1. Append takes a start and
    # a stop page argument. The stop page is not included in the output. The default is the entire file.
    pdf_writer.append(input_file1)
    pdf_writer.append(input_file2, pages=(0, 1))

    if verbose:
        print(f'Writing {pdf_output_file}')
    pdf_writer.write(pdf_output_file)
    pdf_output_file.close()


def convertcsv(input_file, output_file, template_file, verbose, debug):
    # Import a CSV file that has been exported from this code. Read file, everything except the start and end PDF
    # pages, then spit them out into an Excel workbook.
    #
    # This new code can take a template file and assign column numbers based on its headings. Defaults are the
    # old hardcoded columns.
    # TODO: Update for all possible metadata fields?

    title_col = 0
    page_col = 0
    author_col = []

    if template_file:
        if verbose:
            print(f'Using template file {template_file}')
        # Open template file and find headers
        imp_wb = load_workbook(filename=template_file, data_only=True)
        imp_ws = imp_wb.active

        headers = []
        for c in range(1, imp_ws.max_column+1):
            headers.append(imp_ws.cell(row=1, column=c).internal_value)
        for c in range(0, len(headers)):
            if headers[c] == 'title':
                title_col = c + 1
            if headers[c] == 'fpage':
                page_col = c + 1
            for n in range(0, 4):
                if headers[c] == 'author' + str(n+1) + '_fname':
                    author_col.append(c + 1)
        if debug:
            print(author_col)
    else:
        title_col = 1
        page_col = 37
        author_col = [5, 12, 19, 26]

    if debug:
        print(f'{title_col}, {page_col}, {author_col}')

    wb = Workbook()
    ws = wb.active
    wbrow = 1
    dest_filename = output_file + '.xlsx'

    if verbose:
        print(f'Reading {input_file}')

    # Step through each row of input CSV file. Read in each row and assign to variables. Write to Excel file.
    # The columns are hard-coded to correspond to the Digital Commons import columns for Buffalo Law Review.
    # To be more useful, this could write to an output and write the values in the columns based on the
    # headers.
    with open(input_file, newline='') as csvfile:
        data_reader = csv.reader(csvfile, delimiter=",", quotechar='"', quoting=csv.QUOTE_MINIMAL)

        # Check headers to get column numbers; Write headers in destination file
        headers = next(data_reader)
        title_in_col = ''
        page_in_col = ''
        author_in_col = []

        for c in range(0, len(headers)):
            if headers[c] == 'title':
                title_in_col = c
                ws.cell(row=wbrow, column=title_col, value='title')
                title_flag = 1
            if headers[c] == 'start_page':
                page_in_col = c
                ws.cell(row=wbrow, column=page_col, value='fpage')
            for n in range(0, 4):
                if headers[c] == 'f_name' + str(n+1):
                    author_in_col.append(c)
                    ws.cell(row=wbrow, column=author_col[n], value='author' + str(n+1) + '_fname')
                    ws.cell(row=wbrow, column=author_col[n] + 1, value='author' + str(n+1) + '_mname')
                    ws.cell(row=wbrow, column=author_col[n] + 2, value='author' + str(n+1) + '_lname')
                    ws.cell(row=wbrow, column=author_col[n] + 3, value='author' + str(n+1) + '_suffix')
                    ws.cell(row=wbrow, column=author_col[n] + 6, value='author' + str(n+1) + '_is_corporate')
        wbrow += 1

        if debug:
            print(headers)
            print(f'{title_in_col}, {page_in_col}, {author_in_col}')

        for row in data_reader:
            if title_flag:
                ws.cell(row=wbrow, column=title_col, value=row[title_in_col])
            if page_in_col:
                ws.cell(row=wbrow, column=page_col, value=row[page_in_col])
            for n in range(0, 4):
                if author_in_col[n]:
                    try:
                        ws.cell(row=wbrow, column=author_col[n], value=row[author_in_col[n]])
                        ws.cell(row=wbrow, column=author_col[n] + 1, value=row[author_in_col[n]+1])
                        ws.cell(row=wbrow, column=author_col[n] + 2, value=row[author_in_col[n]+2])
                        ws.cell(row=wbrow, column=author_col[n] + 3, value=row[author_in_col[n]+3])
                    # Figure out how to get a row number out of this.
                    except IndexError:
                        print(f'Warning: Potentially missing data in row {row}')
                    if row[author_in_col[n]] and row[author_in_col[n]+2]:
                        ws.cell(row=wbrow, column=author_col[n]+6, value='FALSE')
                    elif row[author_in_col[n]]:
                        ws.cell(row=wbrow, column=author_col[n]+6, value='TRUE')

            wbrow += 1

    csvfile.close()
    wb.save(filename=dest_filename)
    if verbose:
        print(f'Saving {dest_filename}')
    wb.close()


def dirshift(path, verbose, debug, test):
    # File finder for shiftpage. Allows user to drop all files needing a shifted page and the files containing those
    # pages into one directory and automatically shift the pages.
    # TODO: Update to work with any filenames?

    for file in os.listdir(path):
        # Separate filename from path. Retrieve parts of filename as match groups. This is set to work with the standard
        # Hein filenames in pattern **_##JournalAbbrev^^^(%%%%-%%%%) *=item number, # = volume number, ^=start page,
        # % = year(s).
        filename = os.path.basename(file)
        fileparts = re.match(r'(\d\d)_(\d{1,2})([A-Za-z]+)([xvi\d\[\]]+)(\(\d{4}-?\d{0,4}\)).pdf', filename)
        output_file, output_extension = os.path.splitext(file)
        output_file = output_file + "-NEW" + output_extension
        if debug:
            print(f'{filename}, {fileparts}')
        # If there is a match, add one to the item number, then reassemble the file name with a wildcard for the page
        # number. Look for a file that matches this file name. If this file is found, then assign it as the second
        # file and send both to shiftpage.
        if fileparts:
            item_number = int(fileparts.group(1))
            item_number += 1
            if item_number > 9:
                item_number = str(item_number)
            else:
                item_number = "0" + str(item_number)
            filetest = item_number + '_' + fileparts.group(2) + fileparts.group(3) + '*' + fileparts.group(5) + '.pdf'
            print(filetest)
            for f in os.listdir(path):
                if fnmatch.fnmatch(f, filetest):
                    second_file = f
                    if test:
                        print(f'First file: {file}; Second file: {second_file}')
                    else:
                        shiftpage(os.path.join(path, file), os.path.join(path, second_file),
                                  os.path.join(path, output_file), verbose)


def getfilenames(input_file, debug):
    # This collects file names for combinepdf. Like dirshift, it works on standard Hein file names.
    # It takes an input file (the first file), then looks for a file with an item number (the first part of the
    # filename) one higher. If it finds one, it adds it to a list. Then it looks for one higher than that. Once
    # there are no more, it passes the list to combinepdf.

    file_names = []
    input_path, input_filename = os.path.split(input_file)

    # Assign input filename parts to match groups. This is set to work with the standard
    # Hein file names in pattern **_##JournalAbbrev^^^(%%%%-%%%%) *=item number, # = volume number, ^=start page,
    # % = year(s).
    file_parts = re.match(r'(\d\d)_(\d{1,2})([A-Za-z]+)([xvi\d\[\]]+)(\(\d{4}-?\d{0,4}\)).pdf', input_filename)
    if debug:
        print(f'{input_filename}, {file_parts}')
        for r in range(0, file_parts.lastindex):
            print(file_parts.group(r))
    # Loop through each file in the directory looking for a match. If there's a match, add to file_names list
    if file_parts:
        # Set flag. Set the item number to the number from the input file.
        newfile = False
        item_number = int(file_parts.group(1))
        # Look for a series of files where the item number in the file name increases by one, and the rest of the
        # file name is the same (except the page number).
        while item_number:
            # Add a leading zero to single digit item numbers, to match Hein's naming convention.
            if item_number < 10:
                item = '0' + str(item_number)
            else:
                item = str(item_number)
            if debug:
                print(f'file_parts: {file_parts.lastindex}')
            # Make sure that the file name fits the pattern and that there are five parts. If there are,
            # build a test file name that starts with the current item value and has a wildcard for the
            # page number. If not, assign null value to file_test.
            # Right now the journal name is hard coded. The regular expression needs to be tweaked to keep from pulling
            # roman numeral page numbers as part of the journal name. Or at least it needs to filter them after
            # the match groups are set.
            if file_parts.lastindex == 5:
                file_test = item + '_' + file_parts.group(2) + 'BuffLRev' + '*' + file_parts.group(5) + '.pdf'
                print(file_test)
            else:
                print('No matching files found.')
                file_test = ''
            # Step through directory pulled from the input file. Check each file against the test file name. If there's
            # a match, append the name to file_names list. Set the newfile flag to true, so the routine knows that
            # at least one file was found.
            for f in os.listdir(input_path):
                if fnmatch.fnmatch(f, file_test):
                    file_names.append(os.path.join(input_path, f))
                    newfile = True
            # After directory checked, if there was a file match, add one to the item number. If there was no match,
            # set the item number to zero, so the loop ends.
            if newfile:
                item_number += 1
                newfile = False
            else:
                item_number = 0

        return file_names


def capitalize_title(title):
    # Change case of common prepositions and conjunctions to lower case for more accurate headline style capitalization
    title = re.sub(r'(?<!:.)\sAnd\s', ' and ', title)
    title = re.sub(r'(?<!:.)\sBut\s', ' but ', title)
    title = re.sub(r'(?<!:.)\sOf\s', ' of ', title)
    title = re.sub(r'(?<!:.)\sFor\s', ' for ', title)
    title = re.sub(r'(?<!:.)\sOr\s', ' or ', title)
    title = re.sub(r'(?<!:.)\sNor\s', ' nor ', title)
    title = re.sub(r'(?<!:.)\sA\s', ' A ', title)
    title = re.sub(r'(?<!:.)\sAn\s', ' an ', title)
    title = re.sub(r'(?<!:.)\sThe\s', ' the ', title)
    title = re.sub(r'(?<!:.)\sTo\s', ' to ', title)
    title = re.sub(r'(?<!:.)\sAs\s', ' as ', title)
    title = re.sub(r'(?<!:.)\sIn\s', ' in ', title)
    title = re.sub(r'(?<!:.)\sWith\s', ' with ', title)
    title = re.sub(r'(?<!:.)\sAt\s', ' at ', title)
    title = re.sub(r'(?<!:.)\sFrom\s', ' from ', title)
    title = re.sub(r'(?<!:.)\sInto\s', ' into ', title)
    title = re.sub(r'(?<!:.)\sOn\s', ' on ', title)
    title = re.sub(r'(?<!:.)\sIn\s', ' in ', title)
    title = re.sub(r'(?<!:.)\sBy\s', ' by ', title)
    title = re.sub(r'(?<!:.)\sUp\s', ' up ', title)
    title = re.sub(r'(?<!:.)\sOut\s', ' out ', title)
    # Fix hyphenated words and 'S
    title = re.sub(r'- ([A-Z])', lambda x: x.group(1).lower(), title)
    title = re.sub(r'\'S', '\'s', title)
    return title


def dateconvert(year, month=''):
    month = month.capitalize()
    if month == 'January':
        month = '01'
    elif month == "February":
        month = '02'
    elif month == "March":
        month = '03'
    elif month == "April":
        month = '04'
    elif month == "May":
        month = '05'
    elif month == 'June':
        month = '06'
    elif month == 'July':
        month = '07'
    elif month == 'August':
        month = '08'
    elif month == 'September':
        month = '09'
    elif month == 'October':
        month = '10'
    elif month == 'November':
        month = '11'
    elif month == "December":
        month = '12'
    else:
        month = '01'
    date = str(year) + '-' + month + '-01'

    return date
